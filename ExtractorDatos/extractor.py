from openpyxl import load_workbook
import string
from models.hotel import *


EXCLUSIONES = [
    "season rates", "(per room)", "closing agreement",
    "rates includes", "promotion", "not included"
]

def cargar_excel(path_excel, max_row=200) -> DatosExcel:
    wb = load_workbook(path_excel)
    ws = wb.active  # type: ignore

    hoteles: list[HotelExcel] = []
    hotel_actual: HotelExcel | None = None
    tipo_actual: TipoHabitacionExcel | None = None

    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=max_row)):  # type: ignore
        if not any(row):
            continue  # fila vacía

        if row[0] is None:
            continue  # sin nombre

        nombre_raw = str(row[0]).strip()
        nombre_norm = nombre_raw.lower()

        # 🔹 Saltar exclusiones
        if any(nombre_norm.startswith(excl) for excl in EXCLUSIONES):
            continue

        # 🔹 Detectar nuevo hotel (empieza con "hotel" o patrón similar)
        if nombre_raw.endswith("(A)"):
            hotel_actual = HotelExcel(nombre=nombre_raw, tipos=[], habitaciones_directas=[])
            hoteles.append(hotel_actual)
            tipo_actual = None
            continue

        # 🔹 Detectar tipo de habitación (mayúsculas + sin precio en la fila)
        if nombre_raw.isupper() and row[2] is None:
            tipo_actual = TipoHabitacionExcel(nombre=nombre_raw, habitaciones=[])
            if hotel_actual:
                hotel_actual.tipos.append(tipo_actual)
            continue

        # 🔹 Habitaciones (con o sin tipo)
        col_precio = row[2]
        precio_raw = None
        if col_precio is not None and str(col_precio).strip() != "":
            precio_raw = str(col_precio)
            
        habitacion = HabitacionExcel(
            nombre=nombre_raw,
            precio_raw=precio_raw,
            row_idx=i
        )

        if hotel_actual:
            if tipo_actual:
                tipo_actual.habitaciones.append(habitacion)
            else:
                hotel_actual.habitaciones_directas.append(habitacion)

    return DatosExcel(hoteles=hoteles)
